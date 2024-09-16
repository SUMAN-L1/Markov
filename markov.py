import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from transitions import Machine

# Function to create transition matrix
def create_transition_matrix(data):
    unique_states = pd.unique(data.values.ravel())  # Get unique states from the data
    transition_matrix = pd.DataFrame(0, index=unique_states, columns=unique_states)
    
    # Loop through each row and count transitions between consecutive columns
    for i in range(len(data) - 1):
        current_row = data.iloc[i].values
        next_row = data.iloc[i + 1].values
        for current_state, next_state in zip(current_row, next_row):
            if current_state in unique_states and next_state in unique_states:
                transition_matrix.loc[current_state, next_state] += 1
    
    # Normalize by row to get probabilities
    transition_matrix = transition_matrix.div(transition_matrix.sum(axis=1), axis=0)
    transition_matrix.fillna(0, inplace=True)  # Replace NaNs with zeros
    return transition_matrix

# Function to plot heatmap
def plot_heatmap(matrix):
    plt.figure(figsize=(10, 8))
    sns.heatmap(matrix, annot=True, fmt=".4f", cmap="coolwarm")
    plt.title("Transition Matrix Heatmap")
    plt.xlabel("To State")
    plt.ylabel("From State")
    plt.xticks(rotation=45, ha='right', fontsize=12, fontweight='bold')
    plt.yticks(fontsize=12, fontweight='bold')
    plt.tight_layout()
    plt.savefig("transition_matrix_heatmap.png")
    st.pyplot(plt)

# Function to export results to Excel
def export_to_excel(transition_matrix):
    wb = Workbook()
    ws = wb.active
    ws.title = "Transition Matrix"
    
    # Write transition matrix to Excel
    for i, row in enumerate(transition_matrix.index):
        ws.cell(row=i+2, column=1, value=row)
        for j, col in enumerate(transition_matrix.columns):
            ws.cell(row=1, column=j+2, value=col)
            ws.cell(row=i+2, column=j+2, value=transition_matrix.iloc[i, j])
    
    # Add heatmap image
    img = Image("transition_matrix_heatmap.png")
    ws.add_image(img, "J1")
    
    wb.save("Markov_Chain_Results.xlsx")

# Streamlit app
st.title("Markov Chain Analysis by Suman_Econ")

# File upload
uploaded_file = st.file_uploader("Choose an Excel file", type="xlsx")

if uploaded_file:
    data = pd.read_excel(uploaded_file)
    
    # Automatically exclude 'Year' and 'Total' columns
    columns = [col for col in data.columns if col.lower() not in ['year', 'years', 'total']]
    st.write(f"Columns considered for analysis: {columns}")
    
    if columns:
        data_selected = data[columns]
        
        # Ensure data is of the correct type (discrete or categorical)
        for col in data_selected.columns:
            if data_selected[col].dtype == 'float64':
                st.warning(f"Column '{col}' contains numerical values. It may need to be categorized or excluded.")
        
        # Create transition matrix
        transition_matrix = create_transition_matrix(data_selected)
        
        # Plot heatmap
        plot_heatmap(transition_matrix)
        
        # Export results to Excel
        export_to_excel(transition_matrix)
        
        st.success("Analysis complete! Results exported to 'Markov_Chain_Results.xlsx'.")
        
        # Interpretation
        st.subheader("Interpretation")
        st.write("""
        The transition matrix heatmap shows the probabilities of transitioning from one state to another. 
        The rows represent the current state, and the columns represent the next state. 
        Higher values indicate a higher probability of transitioning to that state.
        
        The Markov chain object created can be used to predict future states based on the current state.
        For example, if you start in a particular state, you can use the transition matrix to determine the probability of moving to other states in the next time period.
        """)
